import datetime as dt
import os

import PyPDF2 as pdf
from PIL import Image
import pytesseract
from openpyxl import load_workbook


from temp.decoradores_jocimar import imprimeNome

@imprimeNome
def salvando_escala(escala=""):
    pass


def exibircao_servico_dia() -> list:
    wb = load_workbook(filename="servicos_jo_nov.xlsx")
    planilha_svc = wb.worksheets[0]
    celula_svc = 1
    dia_hoje = dt.date.today().day
    for i in range(6, 37):
        valor = planilha_svc.cell(row=i, column=4).value
        if isinstance(valor, str):
            dia = valor.split("/")[0]
            data = dia_hoje
            if dia_hoje < 10:
                data = f"0{dia_hoje}"
            if dia == str(data):
                celula_svc = i
                break
    servico = planilha_svc.cell(row=celula_svc, column=6).value
    if servico != 'FOLGA':
        servico = servico.split("/")
        tupla_svc = [servico[0], servico[1]]
        return tupla_svc
    else:
        return ["FOLGA"]



salvando_escala()

