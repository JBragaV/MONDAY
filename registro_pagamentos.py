from openpyxl import load_workbook
from datetime import datetime
"""
inicio planilha: Coluna: (3,C)
                 Linha: 3
"""


def _abre_planilha(nome):
    wb = load_workbook(filename=nome)
    planilha = wb["pagamentos"]
    return wb, planilha


def _seleciona_mes(planilha):
    mes_sistema = datetime.now().month
    for indice in range(4, (4+12)):
        mes_planilha = planilha.cell(row=4, column=indice).value.month
        if mes_sistema == mes_planilha:
            return indice


def _cria_lista_contas(planilha, lista, pesquisa=False, mes=0):
    for linha in range(5, 100):
        nome = planilha.cell(row=linha, column=3).value
        if type(nome) == str:
            if pesquisa:
                condicao_conta_mes = planilha.cell(row=linha, column=mes).value
                if condicao_conta_mes.lower() == "não":
                    lista.append(nome)
            else:
                lista.append(nome)
        else:
            break
    return lista


def _seleciona_conta(nome_conta, planilha):
    indice_conta = 0
    lista_contas = []
    lista_contas = _cria_lista_contas(planilha, lista_contas)
    for indice, item in enumerate(lista_contas):
        conta = item.lower().split(" ")
        conta_a_ser_buscada = nome_conta.lower().split("_")
        for i in conta_a_ser_buscada:
            if i in conta:
                indice_conta = indice
                break
    return indice_conta+5


def atualiza_pagamentos(nome_conta):
    wb, planilha = _abre_planilha("pagamentos_meses_22.xlsx")
    indice_mes = _seleciona_mes(planilha)
    indice_conta = _seleciona_conta(nome_conta, planilha)
    planilha.cell(row=indice_conta, column=indice_mes, value="Sim")
    # print(planilha.dimensions)
    wb.save("pagamentos_meses_22.xlsx")


def dizer_contas_a_pagar():
    _, planilha = _abre_planilha("pagamentos_meses_22.xlsx")
    lista_contas_em_aberto = []
    indice_mes = _seleciona_mes(planilha)
    lista_contas_em_aberto = _cria_lista_contas(planilha, lista_contas_em_aberto, True, indice_mes)
    return lista_contas_em_aberto
