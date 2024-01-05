import time
import datetime as dt
import re
import os

import PyPDF2 as pdf
from PIL import Image
import pytesseract
from openpyxl import load_workbook


LISTA_MESES = ['01 - janeiro', "02 - fevereiro", "03 - marco",
               "04 - abril", "05 - maio", "06 - junho",
               "07 - julho", "08 - agosto", "09 - setembro",
               "10 - outubro", "12 - novembro", "12 - dezembro"]


ano = dt.date.today().year
# Abaixo são chaves retiradas do dicinario de pagamentos
# "IscpSociedadeEducacionalLTDA": "faculdade","UNIMED-RIOCOOPTRAB.MED.R.JLTDA": "Unimed","UNIMED-": "Unimed",

DICIONARIO_NOMES_PAGAMENTOS = {"ISCP": "faculdade",
                               "NETSERVICOS": "net",
                               "CLARO": "net",
                               "VIVO": "vivo",
                               "CPFL": "conta_luz",
                               "TRANSFERÊNCIADECONTASALÁRIOPARACONTACORRENTE": "transferencia_salario",
                               "Deal": "aluguel",
                               "Portela": "caixinha",
                               "caixinha@torresorocaba.com.br": "caixinha",
                               "Franca": "fanca",
                               "JocimarBB": "tranferencia_itau",
                               "Bbeu": "tranferencia_itau",
                               "documentofaturadocartao": "cartao_nu_bank",
                               "faturadocartaéonubank": "cartao_nu_bank",
                               "NuBank": "tranferencia_nubank",
                               "Renanda": "pix_renanda",
                               "Unimed": "unimed",
                               "LOLLA": "compras_lola",
                               "SIMPLESNACIONAL": "mei"}


def define_nome(texto):
    for nome in DICIONARIO_NOMES_PAGAMENTOS.keys():
        padrao_nome = f"{nome}".lower()
        if re.findall(padrao_nome, texto.lower(), re.IGNORECASE):
            nome_recibo = DICIONARIO_NOMES_PAGAMENTOS.get(nome)
            return nome_recibo

    return 'recibo_fora_da_listagem'


def define_data(texto):
    texto_novo = str(texto)
    padrao_data = r"\d{2}\/\d{2}\/\d{2,4}"
    datas = re.findall(padrao_data, texto_novo)
    if len(datas) > 0:
        datas = str(datas[0]).split("/")[1]
        mes = LISTA_MESES[int(datas) - 1]
        return mes
    else:
        padrao_data = r"\d{2}\D{3,9}\d{2,4}"
        datas = re.findall(padrao_data, texto_novo)
        if len(datas) > 0:
            padrao_data = r"\D{3,9}"
            meses = re.findall(padrao_data, datas[0])
            mes = str(meses[0]).lower()
            if len(mes) == 3:
                for mes_lista in LISTA_MESES:
                    if mes in mes_lista:
                        mes = mes_lista
                        return mes
            return mes


def cria_pastas(mes, nome, extensao):
    if not os.path.exists("recibos"):
        os.mkdir("recibos")
    if not os.path.exists(os.path.join("recibos", str(ano))):
        os.mkdir(os.path.join("recibos", str(ano)))
    if not os.path.exists(os.path.join("recibos", str(ano), mes)):
        os.mkdir(os.path.join(os.path.join("recibos", str(ano), mes)))
    if not os.path.exists(os.path.join("recibos", str(ano), mes, f"{nome}.{extensao}")):
        caminho_arquivo = os.path.join("recibos", str(ano), mes, f"{nome}.{extensao}")
    else:
        contador = 0
        for comparador in os.listdir(os.path.join("recibos", str(ano), mes)):
            if nome in comparador:
                contador += 1
        caminho_arquivo = os.path.join("recibos", str(ano), mes, f"{nome}_{contador}.{extensao}")
    return caminho_arquivo


def extrair_texto_pdf(caminho_pdf):
    with open(caminho_pdf, "rb") as f:
        texto = pdf.PdfReader(f).pages[0].extract_text()
        texto = re.sub(" ", "", texto)
        mes = define_data(texto)
        nome_recibo = define_nome(texto)
        nome_arquivo = f"{nome_recibo}"
    return mes, nome_arquivo


def tratar_imagem():
    pytesseract.pytesseract.tesseract_cmd = (r"C:\Users\jbrag\OneDrive\Área de Trabalho\Projetos\Pessoal"
                                             r"\Python\Monday\Tesseract-OCR\tesseract.exe")
    foto = pytesseract.image_to_string(Image.open('temp_img.jpg'))
    foto = re.sub(" ", "", foto)
    nome_foto = define_nome(foto)
    mes = define_data(foto)
    return mes, nome_foto


def exibircao_servico_dia():
    wb = load_workbook(filename="servicos_jo_nov.xlsx")
    planilha_svc = wb.worksheets[0]
    celula_svc = 1
    dia_hoje = dt.date.today().day
    for i in range(6, 37):
        valor = planilha_svc.cell(row=i, column=4).value
        if isinstance(valor, str):
            dia = valor.split("/")[0]
            data = dia_hoje
            if dia_hoje < 10:
                data = f"0{dia_hoje}"
            if dia == str(data):
                celula_svc = i
                break
    servico = planilha_svc.cell(row=celula_svc, column=6).value
    if servico != 'FOLGA':
        servico = servico.split("/")
        tupla_svc = [servico[0], servico[1]]
        return tupla_svc
    else:
        return ["FOLGA"]


if __name__ == "__main__":
    exibircao_servico_dia()
